from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from src.importing.ports.services.tabular_reader_port import (
    TabularDocument,
    TabularReaderPort,
    TabularRow,
)


class XlsxTabularReader(TabularReaderPort):
    def can_read(self, *, filename: str | None, content_type: str | None) -> bool:
        name = (filename or "").lower()
        ct = (content_type or "").lower()
        return (
            name.endswith(".xlsx")
            or "spreadsheetml" in ct
            or "application/vnd.ms-excel" in ct
        )

    def read(
        self,
        *,
        filename: str | None,
        content_type: str | None,
        content: bytes,
    ) -> TabularDocument:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        # header row is 1
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            return TabularDocument(headers=[], rows=[])

        headers: list[str] = []
        for cell in header_cells:
            if cell is None:
                headers.append("")
            else:
                headers.append(str(cell).strip())

        # drop empties but keep order for mapping
        normalized_headers = [h for h in headers if h]

        if not normalized_headers:
            return TabularDocument(headers=[], rows=[])

        def _iter_rows():
            # first data row is 2
            for r_idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                # build dict by header position
                values: dict[str, Any] = {}
                empty = True
                for i, header in enumerate(headers):
                    if not header:
                        continue
                    val = row[i] if i < len(row) else None
                    if val is not None and str(val).strip() != "":
                        empty = False
                    values[header] = val
                if empty:
                    continue
                yield TabularRow(row_number=r_idx, values=values)

        return TabularDocument(headers=normalized_headers, rows=_iter_rows())
